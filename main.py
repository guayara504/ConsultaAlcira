import mysql.connector as sql
#from consultas import municipios
from openpyxl import Workbook


conexion = sql.connect(
            host = "192.168.0.19", 
            user = "jfsarmiento", 
            passwd = "Jfs_red07$14",
            database = "zortekv3")


                   
def consultaBuenaventura():
    
    consulta =   """SELECT z01_radicacion_juzgado, z01_radicacion_z01_radicacion, demandante, demandado, clase_proceso, fecha_notificacion, ciudad 
                        FROM z04_estado 
                        WHERE fecha_notificacion = CURDATE() AND ciudad = "BUENAVENTURA"
                        ORDER BY z01_radicacion_juzgado """

    try:
        print("-"*15)
        print("Consultando: Buenaventura")
        cursor = conexion.cursor()
        cursor.execute(consulta)
        consultasBD  = [item for item in cursor.fetchall()]
        print("Finalizado: Buenaventura")
        print("-"*15)
    except:
        print("\nNo Hubo conexion con la base de datos...")
    
    
    
    
    wb = Workbook()
    ruta = 'Buenaventura.xlsx'
    
    hoja = wb.active
    hoja.title = "Buenaventura"
    
    fila = 2 #Fila donde empezamos
    
    hoja["A1"] = "JUZGADO"
    hoja["B1"] = "RADICADO"
    hoja["C1"] = "DEMANDANTE"
    hoja["D1"] = "DEMANDADO"
    hoja["E1"] = "ACTUACION"
    hoja["F1"] = "FECHA"
    hoja["G1"] = "CIUDAD"
    
    for z01_radicacion_juzgado, z01_radicacion_z01_radicacion, demandante, demandado, clase_proceso, fecha_notificacion, ciudad in consultasBD:
        hoja.cell(column=1, row=fila, value=z01_radicacion_juzgado)
        hoja.cell(column=2, row=fila, value=z01_radicacion_z01_radicacion)
        hoja.cell(column=3, row=fila, value=demandante)
        hoja.cell(column=4, row=fila, value=demandado)
        hoja.cell(column=5, row=fila, value=clase_proceso)
        hoja.cell(column=6, row=fila, value=fecha_notificacion)
        hoja.cell(column=7, row=fila, value=ciudad)
        
        fila+=1
    
    wb.save(filename = ruta)

def consultaCartago():
    
    consulta =   """SELECT z01_radicacion_juzgado, z01_radicacion_z01_radicacion, demandante, demandado, clase_proceso, fecha_notificacion, ciudad 
                        FROM z04_estado 
                        WHERE fecha_notificacion = CURDATE() AND ciudad = "CARTAGO"
                        ORDER BY z01_radicacion_juzgado """

    try:
        print("-"*15)
        print("Consultando: Cartago")
        cursor = conexion.cursor()
        cursor.execute(consulta)
        consultasBD  = [item for item in cursor.fetchall()]
        print("Finalizado: Cartago")
        print("-"*15)
    except:
        print("\nNo Hubo conexion con la base de datos...")
    
    
    
    
    wb = Workbook()
    ruta = 'Cartago.xlsx'
    
    hoja = wb.active
    hoja.title = "Cartago"
    
    fila = 2 #Fila donde empezamos
    
    hoja["A1"] = "JUZGADO"
    hoja["B1"] = "RADICADO"
    hoja["C1"] = "DEMANDANTE"
    hoja["D1"] = "DEMANDADO"
    hoja["E1"] = "ACTUACION"
    hoja["F1"] = "FECHA"
    hoja["G1"] = "CIUDAD"
    
    for z01_radicacion_juzgado, z01_radicacion_z01_radicacion, demandante, demandado, clase_proceso, fecha_notificacion, ciudad in consultasBD:
        hoja.cell(column=1, row=fila, value=z01_radicacion_juzgado)
        hoja.cell(column=2, row=fila, value=z01_radicacion_z01_radicacion)
        hoja.cell(column=3, row=fila, value=demandante)
        hoja.cell(column=4, row=fila, value=demandado)
        hoja.cell(column=5, row=fila, value=clase_proceso)
        hoja.cell(column=6, row=fila, value=fecha_notificacion)
        hoja.cell(column=7, row=fila, value=ciudad)
        
        fila+=1
    
    wb.save(filename = ruta)

def consultaBuga():
    
    consulta =   """SELECT z01_radicacion_juzgado, z01_radicacion_z01_radicacion, demandante, demandado, clase_proceso, fecha_notificacion, ciudad 
                        FROM z04_estado 
                        WHERE fecha_notificacion = CURDATE() AND ciudad = "BUGA"
                        ORDER BY z01_radicacion_juzgado """

    try:
        print("-"*15)
        print("Consultando: Buga")
        cursor = conexion.cursor()
        cursor.execute(consulta)
        consultasBD  = [item for item in cursor.fetchall()]
        print("Finalizado: Buga")
        print("-"*15)
    except:
        print("\nNo Hubo conexion con la base de datos...")
    
    
    
    
    wb = Workbook()
    ruta = 'Buga.xlsx'
    
    hoja = wb.active
    hoja.title = "Buga"
    
    fila = 2 #Fila donde empezamos
    
    hoja["A1"] = "JUZGADO"
    hoja["B1"] = "RADICADO"
    hoja["C1"] = "DEMANDANTE"
    hoja["D1"] = "DEMANDADO"
    hoja["E1"] = "ACTUACION"
    hoja["F1"] = "FECHA"
    hoja["G1"] = "CIUDAD"
    
    for z01_radicacion_juzgado, z01_radicacion_z01_radicacion, demandante, demandado, clase_proceso, fecha_notificacion, ciudad in consultasBD:
        hoja.cell(column=1, row=fila, value=z01_radicacion_juzgado)
        hoja.cell(column=2, row=fila, value=z01_radicacion_z01_radicacion)
        hoja.cell(column=3, row=fila, value=demandante)
        hoja.cell(column=4, row=fila, value=demandado)
        hoja.cell(column=5, row=fila, value=clase_proceso)
        hoja.cell(column=6, row=fila, value=fecha_notificacion)
        hoja.cell(column=7, row=fila, value=ciudad)
        
        fila+=1
    
    wb.save(filename = ruta)

def consultaTulua():
    
    consulta =   """SELECT z01_radicacion_juzgado, z01_radicacion_z01_radicacion, demandante, demandado, clase_proceso, fecha_notificacion, ciudad 
                        FROM z04_estado 
                        WHERE fecha_notificacion = CURDATE() AND ciudad = "TULUA"
                        ORDER BY z01_radicacion_juzgado """

    try:
        print("-"*15)
        print("Consultando: Tulua")
        cursor = conexion.cursor()
        cursor.execute(consulta)
        consultasBD  = [item for item in cursor.fetchall()]
        print("Finalizado: Tulua")
        print("-"*15)
    except:
        print("\nNo Hubo conexion con la base de datos...")
    
    
    
    
    wb = Workbook()
    ruta = 'Tulua.xlsx'
    
    hoja = wb.active
    hoja.title = "Tulua"
    
    fila = 2 #Fila donde empezamos
    
    hoja["A1"] = "JUZGADO"
    hoja["B1"] = "RADICADO"
    hoja["C1"] = "DEMANDANTE"
    hoja["D1"] = "DEMANDADO"
    hoja["E1"] = "ACTUACION"
    hoja["F1"] = "FECHA"
    hoja["G1"] = "CIUDAD"
    
    for z01_radicacion_juzgado, z01_radicacion_z01_radicacion, demandante, demandado, clase_proceso, fecha_notificacion, ciudad in consultasBD:
        hoja.cell(column=1, row=fila, value=z01_radicacion_juzgado)
        hoja.cell(column=2, row=fila, value=z01_radicacion_z01_radicacion)
        hoja.cell(column=3, row=fila, value=demandante)
        hoja.cell(column=4, row=fila, value=demandado)
        hoja.cell(column=5, row=fila, value=clase_proceso)
        hoja.cell(column=6, row=fila, value=fecha_notificacion)
        hoja.cell(column=7, row=fila, value=ciudad)
        
        fila+=1
    
    wb.save(filename = ruta)

def consultaPopayan():
    
    consulta =   """SELECT z01_radicacion_juzgado, z01_radicacion_z01_radicacion, demandante, demandado, clase_proceso, fecha_notificacion, ciudad 
                        FROM z04_estado 
                        WHERE fecha_notificacion = CURDATE() AND ciudad = "POPAYAN"
                        ORDER BY z01_radicacion_juzgado """

    try:
        print("-"*15)
        print("Consultando: Popayan")
        cursor = conexion.cursor()
        cursor.execute(consulta)
        consultasBD  = [item for item in cursor.fetchall()]
        print("Finalizado: Popayan")
        print("-"*15)
    except:
        print("\nNo Hubo conexion con la base de datos...")
    
    
    
    
    wb = Workbook()
    ruta = 'Popayan.xlsx'
    
    hoja = wb.active
    hoja.title = "Popayan"
    
    fila = 2 #Fila donde empezamos
    
    hoja["A1"] = "JUZGADO"
    hoja["B1"] = "RADICADO"
    hoja["C1"] = "DEMANDANTE"
    hoja["D1"] = "DEMANDADO"
    hoja["E1"] = "ACTUACION"
    hoja["F1"] = "FECHA"
    hoja["G1"] = "CIUDAD"
    
    for z01_radicacion_juzgado, z01_radicacion_z01_radicacion, demandante, demandado, clase_proceso, fecha_notificacion, ciudad in consultasBD:
        hoja.cell(column=1, row=fila, value=z01_radicacion_juzgado)
        hoja.cell(column=2, row=fila, value=z01_radicacion_z01_radicacion)
        hoja.cell(column=3, row=fila, value=demandante)
        hoja.cell(column=4, row=fila, value=demandado)
        hoja.cell(column=5, row=fila, value=clase_proceso)
        hoja.cell(column=6, row=fila, value=fecha_notificacion)
        hoja.cell(column=7, row=fila, value=ciudad)
        
        fila+=1
    
    wb.save(filename = ruta)

if __name__ == "__main__":
    consultaBuenaventura()
    consultaBuga()
    consultaCartago()
    consultaPopayan()
    consultaTulua()